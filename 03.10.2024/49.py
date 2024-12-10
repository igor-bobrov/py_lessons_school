import matplotlib.pyplot as plt
import numpy as np
import openpyxl

wb = openpyxl.load_workbook("dict_complete_data.xlsx")
sheet1 = wb["Страны"]
sheet2 = wb["Регионы"]
sheet3 = wb["Города"]
countries = dict()
regions_count = dict()
cities_count = dict()

for i in range(2, sheet1.max_row + 1):
    countries.update({sheet1.cell(i, 1).value: sheet1.cell(i, 3).value})
c = 2
for i in range(3, sheet2.max_row + 1):
    if sheet2.cell(i, 2).value == sheet2.cell(i - 1, 2).value:
        regions_count.update({sheet2.cell(i, 2).value: c}); c+=1
    else: 
        c = 1; regions_count.update({sheet2.cell(i, 2).value: c}); c+=1
c = 2
for i in range(2, sheet3.max_row + 1):
    if sheet3.cell(i, 2).value == sheet3.cell(i - 1, 2).value:
        cities_count.update({sheet3.cell(i, 2).value: c}); c+=1
    else: 
        c = 1; cities_count.update({sheet3.cell(i, 2).value: c}); c+=1

cities_sorted = list(cities_count.items())
regions_sorted = list(regions_count.items())
def swap(arr, a, b):
            arr[a], arr[b] = arr[b], arr[a]
for i in range(0, len(cities_sorted)- 1):
            swaped = False
            for j in range(0, len(cities_sorted) - 1 - i):
                if cities_sorted[j][1] > cities_sorted[j+1][1]:
                    swap(cities_sorted, j, j+1)
                    swaped = True
            if not swaped:
                break
for i in range(0, len(regions_sorted)- 1):
            swaped = False
            for j in range(0, len(regions_sorted) - 1 - i):
                if regions_sorted[j][1] > regions_sorted[j+1][1]:
                    swap(regions_sorted, j, j+1)
                    swaped = True
            if not swaped:
                break
cities_sorted = cities_sorted[-10:]
regions_sorted = regions_sorted[-10:]
x1_axis = []
y1_axis = []
x2_axis = []
y2_axis = []
for i in cities_sorted:
      x1_axis.append(countries[i[0]])
      y1_axis.append(i[1])
for i in regions_sorted:
      x2_axis.append(countries[i[0]])
      y2_axis.append(i[1])
      
fig, (f1,f2) = plt.subplots(1, 2, figsize = (20, 6), gridspec_kw={'wspace':0.3})
f1.set_xticks(np.arange(len(x1_axis)), x1_axis, rotation = 45)
f1.set_xlabel("города")
f1.bar(np.arange(len(x1_axis)), y1_axis, align='center', color = 'c')
f2.set_xticks(np.arange(len(x2_axis)), x2_axis, rotation = 45)
f2.bar(np.arange(len(x2_axis)), y2_axis, align='center', color = 'pink')
f2.set_xlabel("регионы")
plt.subplots_adjust(bottom=0.27)
plt.show()